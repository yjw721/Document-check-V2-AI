# -*- coding: utf-8 -*-
"""
Excel (.xlsx/.xlsm) 低级错误检测引擎
=========================================
覆盖检测项：
    1. 公式报错：#N/A / #VALUE! / #DIV/0! / #REF! / #NAME? / #NULL! / #NUM! 等
    2. 数字文本化存储、单元格格式混乱
    3. 整行 / 整列空白、表头空单元格、表头重复字段
    4. 多余合并单元格、无意义空白工作表
    5. 工作表名称为空 / 默认名 / 含非法特殊字符
    6. 单元格首尾空格、全角空格、隐藏行列

实现要点：
    openpyxl 双次只读加载：
        data_only=True  -> 取缓存计算值，用于发现公式错误值
        data_only=False -> 取公式文本与格式，用于格式类检测
    全部本地计算，不联网、不修改原始文件。
"""

from __future__ import annotations

import os
import re
from typing import Any, Callable, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from checkers.base import (
    STATUS_ISSUE,
    STATUS_PASS,
    STATUS_UNREADABLE,
    FileResult,
    Issue,
    clip,
    excel_location,
    precheck_ooxml,
)
from config.config_manager import RuleConfig
from checkers.textnorm_checker import TextNormChecker
from checkers.fluency_checker import FluencyChecker
from checkers.custom_rules import CustomRuleEngine
from checkers.wordbank import WordBankEngine

FULL_WIDTH_SPACE = "\u3000"
DEFAULT_SHEET_NAME_RE = re.compile(r"^(Sheet|sheet|工作表|表)\s*\d*$")
# 可识别为数字的文本（含千分位、百分号、正负号、科学计数）
NUMERIC_TEXT_RE = re.compile(r"^[+-]?(\d{1,3}(,\d{3})+|\d+)(\.\d+)?%?$")
# 扫描上限，防止超大表卡死界面
MAX_SCAN_ROWS = 3000
MAX_SCAN_COLS = 200


class ExcelChecker:
    """Excel 工作簿检测器。一个实例对应一次文件检测。"""

    def __init__(self, config: RuleConfig, fluency_sensitivity: str = "normal",
                 progress: Optional[Callable[[float, str, str], bool]] = None) -> None:
        self.cfg = config
        self.kind = "excel"
        self.issues: List[Issue] = []
        self._limit = config.max_issues_per_file()
        self._fluency_sensitivity = fluency_sensitivity
        self._progress = progress

    # ------------------------------------------------------------------
    # 进度钩子（阶段百分比 + 思考日志；返回 False 表示任务已取消）
    # ------------------------------------------------------------------
    def _hook(self, pct: float, stage: str, log: str) -> bool:
        if not self._progress:
            return True
        try:
            return bool(self._progress(pct, stage, log))
        except Exception:  # noqa: BLE001 - 钩子异常不影响检测
            return True

    # ------------------------------------------------------------------
    # 内部工具
    # ------------------------------------------------------------------
    def _on(self, rule_key: str) -> bool:
        if len(self.issues) >= self._limit:
            return False
        return self.cfg.is_enabled(self.kind, rule_key)

    def _add(self, rule_key: str, location: str, detail: str, snippet: str = "") -> None:
        if len(self.issues) >= self._limit:
            return
        self.issues.append(
            Issue(
                rule_key=rule_key,
                rule_title=self.cfg.title(self.kind, rule_key),
                severity=self.cfg.severity(self.kind, rule_key),
                location=location,
                detail=detail,
                snippet=clip(snippet, 100),
                suggestion=self.cfg.suggestion(self.kind, rule_key),
            )
        )

    def _note(self, msg: str) -> None:
        self.issues.append(
            Issue(
                rule_key="engine_note",
                rule_title="检测引擎提示",
                severity="low",
                location="—",
                detail=msg,
                snippet="",
                suggestion="该部分结构特殊，建议人工重点复核。",
            )
        )

    # ------------------------------------------------------------------
    # 主入口
    # ------------------------------------------------------------------
    def check(self, path: str, display_name: Optional[str] = None) -> FileResult:
        """检测单个 Excel 文件，异常一律转为「无法解析」结果，不中断批量任务。"""
        name = display_name or os.path.basename(path)
        try:
            size = os.path.getsize(path)
        except OSError:
            size = 0

        result = FileResult(file_name=name, file_path=path, file_type="Excel", file_size=size)

        # 1) 结构预检
        if not self._hook(2.0, "parse", f"正在解析「{name}」：xlsx 结构预检…"):
            return result
        reason = precheck_ooxml(path)
        if reason:
            result.status = STATUS_UNREADABLE
            result.error_message = reason
            return result

        wb_val = None   # data_only=True
        wb_raw = None   # data_only=False
        if not self._hook(5.0, "parse", f"正在解析「{name}」工作簿（缓存值 + 公式）…"):
            return result
        try:
            wb_val = openpyxl.load_workbook(path, data_only=True, read_only=False, keep_links=False)
            wb_raw = openpyxl.load_workbook(path, data_only=False, read_only=False, keep_links=False)
        except Exception as exc:  # noqa: BLE001
            for wb in (wb_val, wb_raw):
                try:
                    if wb is not None:
                        wb.close()
                except Exception:  # noqa: BLE001
                    pass
            msg = clip(str(exc), 120)
            if "encrypt" in msg.lower() or "password" in msg.lower():
                msg = "工作簿已加密，需密码打开，无法解析"
            result.status = STATUS_UNREADABLE
            result.error_message = f"工作簿解析失败：{msg}"
            return result

        self.issues = []
        # 文字规范 / 表述问题 / 自定义规则 / 自定义词库检测引擎（共享同一 issues 列表）
        self.tn = TextNormChecker(self.cfg, self.issues, self._limit)
        self.fl = FluencyChecker(self.cfg, self.issues, self._limit, self._fluency_sensitivity)
        self.cust = CustomRuleEngine(self.issues, self._limit)
        self.wb = WordBankEngine(self.issues, self._limit)
        stat = {"sheets": 0, "cells": 0, "formulas": 0, "merged": 0}

        try:
            stat["sheets"] = len(wb_val.sheetnames)
            if not self._hook(8.0, "page", f"正在定位「{name}」工作表：表名规范检测…"):
                return result
            self._check_sheet_names(wb_val.sheetnames)

            for si, sheet_name in enumerate(wb_val.sheetnames, start=1):
                try:
                    if not self._hook(10.0 + 75.0 * (si - 1) / max(len(wb_val.sheetnames), 1),
                                      "page", f"正在解析工作表「{sheet_name}」：读取行列矩阵…"):
                        return result
                    ws_val = wb_val[sheet_name]
                    ws_raw = wb_raw[sheet_name] if sheet_name in wb_raw.sheetnames else ws_val
                    self._check_sheet(sheet_name, ws_val, ws_raw, stat, si)
                except Exception as exc:  # noqa: BLE001
                    self._note(f"工作表「{sheet_name}」检测异常：{type(exc).__name__}，已跳过该表")
        except Exception as exc:  # noqa: BLE001
            self._note(f"工作簿检测过程异常：{type(exc).__name__}")
        finally:
            for wb in (wb_val, wb_raw):
                try:
                    wb.close()
                except Exception:  # noqa: BLE001
                    pass

        if not self._hook(100.0, "summary", f"正在收集「{name}」位置信息：工作表/行列号…"):
            return result
        result.issues = self.issues
        result.stats = stat
        result.truncated = len(self.issues) >= self._limit
        result.status = STATUS_ISSUE if self.issues else STATUS_PASS
        return result

    # ------------------------------------------------------------------
    # 工作表名称检测
    # ------------------------------------------------------------------
    def _check_sheet_names(self, names: List[str]) -> None:
        if not self._on("sheet_name_invalid"):
            return
        illegal: List[str] = self.cfg.param(self.kind, "sheet_name_invalid", "illegal_chars",
                                            [":", "\\", "/", "?", "*", "[", "]"]) or []
        flag_default = bool(self.cfg.param(self.kind, "sheet_name_invalid", "flag_default_name", True))

        for name in names:
            problems: List[str] = []
            if not name or not name.strip():
                problems.append("名称为空或全为空白")
            if name != name.strip():
                problems.append("名称首尾存在空格")
            hit = [c for c in illegal if c in name]
            if hit:
                problems.append("含非法字符：" + " ".join(hit))
            if FULL_WIDTH_SPACE in name:
                problems.append("含全角空格")
            if flag_default and DEFAULT_SHEET_NAME_RE.match(name.strip()):
                problems.append("使用系统默认名称，语义不明确")
            if len(name) > 31:
                problems.append(f"名称长度 {len(name)} 超过 Excel 上限 31 字符")
            if problems:
                self._add("sheet_name_invalid", f"工作表「{name}」",
                          "；".join(problems), name)

    # ------------------------------------------------------------------
    # 单表检测
    # ------------------------------------------------------------------
    def _check_sheet(self, sheet_name: str, ws_val: Any, ws_raw: Any, stat: Dict[str, int],
                     si: int = 1) -> None:
        """对单个工作表执行全部检测项。"""
        max_row = min(int(ws_val.max_row or 0), MAX_SCAN_ROWS)
        max_col = min(int(ws_val.max_column or 0), MAX_SCAN_COLS)

        # ---- 读取值矩阵（缓存计算值）与公式矩阵 ----
        values: List[List[Any]] = []
        for row in ws_val.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            values.append(list(row))
        while values and all(v is None or str(v).strip() == "" for v in values[-1]):
            values.pop()   # 去掉尾部纯空行，避免误判「整行空白」

        real_rows = len(values)
        non_empty_cells = sum(1 for r in values for v in r if v is not None and str(v).strip() != "")
        stat["cells"] += non_empty_cells

        # ---- 空白工作表 ----
        if non_empty_cells == 0:
            if self._on("empty_sheet"):
                self._add("empty_sheet", f"工作表「{sheet_name}」",
                          "工作表无任何有效内容（全部单元格为空），属冗余空白表", "")
            return

        # ---- 公式报错 ----
        error_values: List[str] = self.cfg.param(
            self.kind, "formula_error", "error_values",
            ["#N/A", "#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!"]
        ) or []
        err_set = {e.upper() for e in error_values}

        if self._on("formula_error"):
            found_err: List[Tuple[str, str, str]] = []   # (坐标, 错误值, 公式)
            for r_idx in range(1, real_rows + 1):
                for c_idx in range(1, max_col + 1):
                    val = values[r_idx - 1][c_idx - 1] if c_idx - 1 < len(values[r_idx - 1]) else None
                    if val is None:
                        continue
                    text = str(val).strip().upper()
                    if text in err_set:
                        coord = f"{get_column_letter(c_idx)}{r_idx}"
                        formula = ""
                        try:
                            raw = ws_raw.cell(row=r_idx, column=c_idx).value
                            if isinstance(raw, str) and raw.startswith("="):
                                formula = raw
                        except Exception:  # noqa: BLE001
                            pass
                        found_err.append((coord, str(val).strip(), formula))
                        if len(found_err) >= 40:
                            break
                if len(found_err) >= 40:
                    break
            for coord, errv, formula in found_err:
                r_num = int("".join(filter(str.isdigit, coord)) or "0")
                c_letter = "".join(filter(str.isalpha, coord))
                from openpyxl.utils import column_index_from_string
                try:
                    c_num = column_index_from_string(c_letter)
                except Exception:  # noqa: BLE001
                    c_num = 1
                self._add("formula_error", excel_location(sheet_name, r_num, c_num),
                          f"单元格计算结果为错误值 {errv}" + (f"，公式：{clip(formula, 60)}" if formula else ""),
                          formula or errv)

        # ---- 统计公式数量 ----
        try:
            for row in ws_raw.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        stat["formulas"] += 1
        except Exception:  # noqa: BLE001
            pass

        # ---- 表头检测（首行）----
        header = values[0] if values else []
        header_texts = ["" if v is None else str(v).strip() for v in header]
        # 表头有效范围：到最后一个非空表头为止
        last_header = 0
        for i, t in enumerate(header_texts):
            if t:
                last_header = i + 1

        if self._on("empty_header_cell") and last_header > 1:
            empties = [i + 1 for i in range(last_header) if not header_texts[i]]
            if empties:
                cols = "、".join(get_column_letter(i) + " 列" for i in empties[:8])
                self._add("empty_header_cell", f"{sheet_name}!第 1 行",
                          f"表头存在 {len(empties)} 个空单元格（{cols}）",
                          " | ".join(header_texts[:8]))

        if self._on("duplicate_header") and last_header > 1:
            seen: Dict[str, int] = {}
            dup: List[str] = []
            for i in range(last_header):
                t = header_texts[i]
                if not t:
                    continue
                if t in seen:
                    dup.append(f"{t}（{get_column_letter(seen[t])} 列 与 {get_column_letter(i + 1)} 列）")
                else:
                    seen[t] = i + 1
            if dup:
                self._add("duplicate_header", f"{sheet_name}!第 1 行",
                          f"表头存在 {len(dup)} 组重复字段：" + "；".join(dup[:5]),
                          " | ".join(header_texts[:8]))

        # ---- 整行空白（数据区域中间）----
        if self._on("empty_row"):
            blank_rows: List[int] = []
            for r_idx in range(2, real_rows + 1):   # 首行为表头，从第 2 行开始
                row_vals = values[r_idx - 1]
                if all(v is None or str(v).strip() == "" for v in row_vals):
                    blank_rows.append(r_idx)
            if blank_rows:
                shown = "、".join(f"第 {r} 行" for r in blank_rows[:10])
                more = f" 等共 {len(blank_rows)} 行" if len(blank_rows) > 10 else ""
                self._add("empty_row", f"工作表「{sheet_name}」",
                          f"数据区域中存在整行空白：{shown}{more}，会打断数据连续性", "")

        # ---- 整列空白（数据区域中间）----
        if self._on("empty_col") and last_header > 1:
            blank_cols: List[int] = []
            for c_idx in range(1, last_header + 1):
                col_vals = [values[r][c_idx - 1] if c_idx - 1 < len(values[r]) else None
                            for r in range(real_rows)]
                if all(v is None or str(v).strip() == "" for v in col_vals):
                    blank_cols.append(c_idx)
            if blank_cols:
                shown = "、".join(f"{get_column_letter(c)} 列" for c in blank_cols[:10])
                self._add("empty_col", f"工作表「{sheet_name}」",
                          f"数据区域中存在整列空白：{shown}（共 {len(blank_cols)} 列）", "")

        # ---- 数字文本化存储 / 首尾空格 / 全角空格 ----
        self._check_cell_text(sheet_name, ws_raw, values, real_rows, max_col)

        # ---- 文字规范 / 表述问题检测（本地词库 + 正则，离线）----
        self._check_textnorm(sheet_name, values, real_rows, max_col)

        # ---- 单元格格式混乱 ----
        self._check_format_chaos(sheet_name, ws_raw, real_rows, last_header or max_col)

        # ---- 合并单元格 ----
        self._check_merged(sheet_name, ws_val, stat)

        # ---- 隐藏行列 ----
        self._check_hidden(sheet_name, ws_raw)

    # ------------------------------------------------------------------
    # 文本类问题
    # ------------------------------------------------------------------
    def _check_cell_text(self, sheet_name: str, ws_raw: Any, values: List[List[Any]],
                         real_rows: int, max_col: int) -> None:
        """数字文本化、首尾空格、全角空格。"""
        chk_num_text = self._on("number_stored_as_text")
        chk_trail = self._on("trailing_space_cell")
        chk_full = self._on("full_width_space_cell")
        if not (chk_num_text or chk_trail or chk_full):
            return

        num_text_hits: List[Tuple[str, str]] = []
        trail_hits: List[Tuple[str, str]] = []
        full_hits: List[Tuple[str, str]] = []

        for r_idx in range(1, real_rows + 1):
            row = values[r_idx - 1]
            for c_idx in range(1, min(max_col, len(row)) + 1):
                val = row[c_idx - 1]
                if not isinstance(val, str) or val == "":
                    continue
                coord = f"{get_column_letter(c_idx)}{r_idx}"
                loc = excel_location(sheet_name, r_idx, c_idx)

                # 数字文本化：字符串但内容是纯数字，且不是公式结果
                if chk_num_text and len(num_text_hits) < 30 and self._is_numeric_text(val):
                    try:
                        raw = ws_raw.cell(row=r_idx, column=c_idx)
                        is_formula = isinstance(raw.value, str) and raw.value.startswith("=")
                        fmt = str(raw.number_format or "")
                    except Exception:  # noqa: BLE001
                        is_formula, fmt = False, ""
                    # 文本格式(@) 或常规格式下的字符串数字都算问题
                    if not is_formula:
                        num_text_hits.append(
                            (coord, f"{val.strip()}（格式：{fmt if fmt and fmt != 'General' else '常规'}）")
                        )

                if chk_trail and len(trail_hits) < 30:
                    if val != val.strip(" \t\u3000\u00a0"):
                        trail_hits.append((coord, repr(val)[:60]))

                if chk_full and len(full_hits) < 30:
                    if FULL_WIDTH_SPACE in val:
                        full_hits.append((coord, val.replace(FULL_WIDTH_SPACE, "␣")[:60]))

        if num_text_hits:
            shown = "、".join(f"{c}={v}" for c, v in num_text_hits[:8])
            self._add("number_stored_as_text", f"工作表「{sheet_name}」",
                      f"检测到 {len(num_text_hits)} 个「数字以文本形式存储」的单元格：{shown}"
                      + ("…" if len(num_text_hits) > 8 else ""),
                      shown)
        if trail_hits:
            shown = "、".join(c for c, _ in trail_hits[:10])
            self._add("trailing_space_cell", f"工作表「{sheet_name}」",
                      f"共 {len(trail_hits)} 个单元格首尾含多余空格：{shown}"
                      + ("…" if len(trail_hits) > 10 else ""),
                      trail_hits[0][1] if trail_hits else "")
        if full_hits:
            shown = "、".join(c for c, _ in full_hits[:10])
            self._add("full_width_space_cell", f"工作表「{sheet_name}」",
                      f"共 {len(full_hits)} 个单元格含全角空格：{shown}"
                      + ("…" if len(full_hits) > 10 else ""),
                      full_hits[0][1] if full_hits else "")

    # ------------------------------------------------------------------
    # 文字规范 / 表述问题检测（文本单元格）
    # ------------------------------------------------------------------
    def _check_textnorm(self, sheet_name: str, values: List[List[Any]],
                       real_rows: int, max_col: int) -> None:
        """对文本单元格做口语化 / 冗余 / 近义词 / 错别字 / 缩写 / 单位等检测，并接入自定义引擎。"""
        for r_idx in range(1, real_rows + 1):
            if not self._hook(0.0, "format_error",
                              f"正在检测工作表「{sheet_name}」第 {r_idx}/{real_rows} 行单元格（文本规范/通顺度/词库）…"):
                return
            row = values[r_idx - 1]
            for c_idx in range(1, min(max_col, len(row)) + 1):
                val = row[c_idx - 1]
                if not isinstance(val, str) or not val.strip():
                    continue
                loc = excel_location(sheet_name, r_idx, c_idx)
                self.tn.check_text(loc, val)
                self.fl.check_text(loc, val)
                self.cust.check_text("excel", loc, val)
                self.wb.check_text("excel", loc, val)

    @staticmethod
    def _is_numeric_text(val: str) -> bool:
        """
        判断字符串是否属于「数字被存成文本」。

        排除以下合理场景，避免误报：
            - 以 0 开头的编号（如 001、0755），前导零是业务需要
            - 长度 >= 15 的纯数字（身份证、银行卡、长编号），必须用文本存储
        """
        stripped = val.strip()
        if not stripped or not NUMERIC_TEXT_RE.match(stripped):
            return False
        digits = stripped.replace(",", "").replace("%", "").lstrip("+-")
        if digits.startswith("0") and digits not in ("0", "0.0") and not digits.startswith("0."):
            return False
        if len(digits.replace(".", "")) >= 15:
            return False
        return True

    # ------------------------------------------------------------------
    # 格式混乱
    # ------------------------------------------------------------------
    def _check_format_chaos(self, sheet_name: str, ws_raw: Any, real_rows: int, max_col: int) -> None:
        """同列数字格式 / 数据类型不统一。"""
        if not self._on("format_chaos") or real_rows < 3:
            return
        min_samples = int(self.cfg.param(self.kind, "format_chaos", "min_samples", 3) or 3)
        scan_rows = min(real_rows, 500)

        for c_idx in range(1, max_col + 1):
            fmts: Dict[str, int] = {}
            types: Dict[str, int] = {}
            samples = 0
            try:
                for r_idx in range(2, scan_rows + 1):   # 跳过表头
                    cell = ws_raw.cell(row=r_idx, column=c_idx)
                    val = cell.value
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        continue
                    if isinstance(val, str) and val.startswith("="):
                        continue
                    samples += 1
                    fmt = str(cell.number_format or "General")
                    fmts[fmt] = fmts.get(fmt, 0) + 1
                    tname = type(val).__name__
                    if tname in ("int", "float"):
                        tname = "数字"
                    elif tname == "str":
                        tname = "文本"
                    elif tname == "datetime":
                        tname = "日期"
                    elif tname == "bool":
                        tname = "布尔"
                    types[tname] = types.get(tname, 0) + 1
            except Exception:  # noqa: BLE001
                continue

            if samples < min_samples:
                continue

            col_letter = get_column_letter(c_idx)
            # 数据类型混排（文本与 数字/日期 混排）
            if "文本" in types and len(types) >= 2:
                desc = "、".join(f"{k} {v} 个" for k, v in types.items())
                self._add("format_chaos", f"{sheet_name}!{col_letter} 列",
                          f"该列数据类型混排（{desc}），统计与排序易出错", "")
                continue
            # 数字格式不统一（超过 2 种且样本足够）
            if len(fmts) >= 3:
                desc = "、".join(f"{k}({v})" for k, v in list(fmts.items())[:4])
                self._add("format_chaos", f"{sheet_name}!{col_letter} 列",
                          f"该列存在 {len(fmts)} 种不同的单元格数字格式：{desc}", "")

    # ------------------------------------------------------------------
    # 合并单元格
    # ------------------------------------------------------------------
    def _check_merged(self, sheet_name: str, ws: Any, stat: Dict[str, int]) -> None:
        """空内容合并单元格 / 数据区域内合并单元格。"""
        if not self._on("redundant_merged_cells"):
            return
        try:
            ranges = list(ws.merged_cells.ranges)
        except Exception:  # noqa: BLE001
            return
        if not ranges:
            return
        stat["merged"] += len(ranges)

        empty_merges: List[str] = []
        body_merges: List[str] = []
        for rng in ranges:
            try:
                coord = str(rng)
                top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
                if top_left is None or str(top_left).strip() == "":
                    empty_merges.append(coord)
                elif rng.min_row > 1:   # 首行合并常为标题，允许；数据区合并影响筛选
                    body_merges.append(coord)
            except Exception:  # noqa: BLE001
                continue

        if empty_merges:
            self._add("redundant_merged_cells", f"工作表「{sheet_name}」",
                      f"存在 {len(empty_merges)} 处内容为空的合并单元格："
                      + "、".join(empty_merges[:8]) + ("…" if len(empty_merges) > 8 else ""),
                      "")
        if body_merges:
            self._add("redundant_merged_cells", f"工作表「{sheet_name}」",
                      f"数据区域内存在 {len(body_merges)} 处合并单元格，将影响排序、筛选与公式引用："
                      + "、".join(body_merges[:8]) + ("…" if len(body_merges) > 8 else ""),
                      "")

    # ------------------------------------------------------------------
    # 隐藏行列
    # ------------------------------------------------------------------
    def _check_hidden(self, sheet_name: str, ws: Any) -> None:
        if not self._on("hidden_row_col"):
            return
        hidden_rows: List[int] = []
        hidden_cols: List[str] = []
        try:
            for idx, dim in (ws.row_dimensions or {}).items():
                if getattr(dim, "hidden", False):
                    hidden_rows.append(int(idx))
            for key, dim in (ws.column_dimensions or {}).items():
                if getattr(dim, "hidden", False):
                    hidden_cols.append(str(key))
        except Exception:  # noqa: BLE001
            return

        parts: List[str] = []
        if hidden_rows:
            parts.append(f"隐藏行 {len(hidden_rows)} 个（"
                         + "、".join(str(r) for r in sorted(hidden_rows)[:8]) + "）")
        if hidden_cols:
            parts.append(f"隐藏列 {len(hidden_cols)} 个（" + "、".join(sorted(hidden_cols)[:8]) + "）")
        if parts:
            self._add("hidden_row_col", f"工作表「{sheet_name}」",
                      "存在被隐藏的行列：" + "；".join(parts), "")


def check_excel(path: str, config: RuleConfig, display_name: Optional[str] = None,
                fluency_sensitivity: str = "normal",
                progress: Optional[Callable[[float, str, str], bool]] = None) -> FileResult:
    """对外统一函数式入口。"""
    return ExcelChecker(config, fluency_sensitivity, progress).check(path, display_name)
